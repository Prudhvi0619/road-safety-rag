from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from html import escape
from pathlib import Path

from .catalog import get_metric
from .models import RoadContext, RuleStatus, ThresholdResult, ThresholdSet
from .road_context import RoadTypeResolver, median_integer, parse_coordinate_text
from .service import StandardsRAG

NUMBER_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")
AUDIT_METRIC_KEYS = [
    "min_lane_width",
    "min_sign_height",
    "min_kerb_height",
    "min_w_beam_barrier_height",
    "min_concrete_barrier_height",
    "min_radius_curvature",
]
CHECK_METRIC_KEYS = {
    "Lane_Width": "min_lane_width",
    "Radius_of_Curvature": "min_radius_curvature",
    "Kerb_Height": "min_kerb_height",
    "WBeam_Barrier_Height": "min_w_beam_barrier_height",
    "Concrete_Barrier_Height": "min_concrete_barrier_height",
    "Traffic_Sign_Mounting_Height": "min_sign_height",
}


def extract_numbers(value: object) -> list[float]:
    if value is None:
        return []
    try:
        import pandas as pd

        if pd.isna(value):
            return []
    except (ImportError, TypeError, ValueError):
        pass
    return [float(match.group()) for match in NUMBER_RE.finditer(str(value))]


@dataclass(frozen=True)
class CheckResult:
    status: str
    detail: str
    failed: bool = False
    review_required: bool = False


def evaluate_measurements(values: list[float], rule: ThresholdResult) -> CheckResult:
    if not values:
        return CheckResult("NOT_OBSERVED", "Measurement absent")
    rendered = ", ".join(f"{value:g} m" for value in values)
    if not rule.audit_ready:
        candidate = ""
        if rule.value_m is not None and rule.comparator is not None:
            if rule.comparator == "range" and rule.second_value_m is not None:
                low, high = sorted((rule.value_m, rule.second_value_m))
                candidate = f"; evidence-backed candidate {low:g}-{high:g} m"
            else:
                candidate = f"; evidence-backed candidate {rule.comparator} {rule.value_m:g} m"
        return CheckResult(
            "REVIEW_REQUIRED",
            f"{rendered}{candidate}; not used for compliance ({rule.status.value}: {rule.reason})",
            review_required=True,
        )
    assert rule.value_m is not None and rule.comparator is not None
    low = rule.value_m
    high = rule.second_value_m

    if rule.comparator == "=":
        exact = [
            abs(value - low) <= max(1e-6, abs(low) * 1e-6)
            for value in values
        ]
        source = rule.citation.label if rule.citation else "verified source"
        if all(exact):
            return CheckResult(
                "PASS", f"{rendered}; exact specified dimension {low:g} m ({source})"
            )
        return CheckResult(
            "REVIEW_REQUIRED",
            (
                f"{rendered}; specified dimension = {low:g} m ({source}); "
                "a reviewer-approved construction/measurement tolerance is required "
                "before declaring a deviation compliant or non-compliant"
            ),
            review_required=True,
        )

    def passes(value: float) -> bool:
        if rule.comparator == ">=":
            return value >= low
        if rule.comparator == ">":
            return value > low
        if rule.comparator == "<=":
            return value <= low
        if rule.comparator == "<":
            return value < low
        if rule.comparator == "range" and high is not None:
            lower, upper = sorted((low, high))
            return lower <= value <= upper
        return False

    outcomes = [passes(value) for value in values]
    source = rule.citation.label if rule.citation else "verified source"
    requirement = (
        f"{rule.comparator} {low:g} m"
        if rule.comparator != "range"
        else f"between {min(low, high or low):g} and {max(low, high or low):g} m"
    )
    if all(outcomes):
        return CheckResult("PASS", f"{rendered}; requirement {requirement} ({source})")
    return CheckResult(
        "FAIL",
        f"{rendered}; requirement {requirement} ({source})",
        failed=True,
    )


class RoadSafetyAuditPipeline:
    """Combine measured outputs with verified RAG rules; never let prose decide compliance."""

    def __init__(
        self,
        metrics_file_path: str | Path,
        rag: StandardsRAG,
        road_context: RoadContext | None = None,
        allow_network_road_lookup: bool = False,
    ):
        self.metrics_file_path = Path(metrics_file_path).resolve()
        self.rag = rag
        self.road_context_override = road_context
        self.allow_network_road_lookup = allow_network_road_lookup

    def run(
        self, output_path: str | Path, html_output: str | Path | None = None
    ) -> Path:
        import pandas as pd

        frame = (
            pd.read_excel(self.metrics_file_path)
            if self.metrics_file_path.suffix.casefold() in {".xlsx", ".xls"}
            else pd.read_csv(self.metrics_file_path)
        )
        context = self._resolve_context(frame)
        observed_keys = self._observed_metric_keys(frame, context)
        threshold_set = self.rag.extract_all(context, metric_keys=observed_keys)
        for key in AUDIT_METRIC_KEYS:
            if key in threshold_set.results:
                continue
            metric = get_metric(key)
            threshold_set.results[key] = ThresholdResult(
                metric_key=metric.key,
                metric_name=metric.name,
                status=RuleStatus.NOT_APPLICABLE,
                reason=(
                    "No usable measurement was present in the corresponding input column; "
                    "threshold retrieval was not run."
                ),
            )
        audited = self._audit_rows(frame, threshold_set)
        output = Path(output_path).resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        self._write_workbook(audited, threshold_set, output)
        if html_output is not None:
            html_path = Path(html_output).resolve()
            html_path.parent.mkdir(parents=True, exist_ok=True)
            self._write_html(audited, threshold_set, html_path, self.metrics_file_path)
        return output

    @staticmethod
    def _observed_metric_keys(frame, context: RoadContext | None = None) -> list[str]:
        columns = {
            "min_lane_width": "Tracked Lane Widths (m)",
            "min_sign_height": "Height",
            "min_kerb_height": "kerb_height",
            "min_w_beam_barrier_height": "WBeam_crash barrier_height",
            "min_concrete_barrier_height": "Concrete_crash_barrier_height",
            "min_radius_curvature": "Radius of Curvature (m)",
        }
        observed = [
            key
            for key, column in columns.items()
            if column in frame.columns
            and any(extract_numbers(value) for value in frame[column].tolist())
        ]
        generic_barrier = "Crash barrier_height"
        if (
            generic_barrier in frame.columns
            and any(extract_numbers(value) for value in frame[generic_barrier].tolist())
            and context is not None
        ):
            key = {
                "w_beam": "min_w_beam_barrier_height",
                "concrete": "min_concrete_barrier_height",
            }.get(context.barrier_type)
            if key and key not in observed:
                observed.append(key)
        return observed

    def _resolve_context(self, frame) -> RoadContext:
        coordinates = [
            coordinate
            for coordinate in (
                parse_coordinate_text(value) for value in frame.get("Raw OCR Text", [])
            )
            if coordinate is not None
        ]
        rag_settings = getattr(self.rag, "settings", None)
        cache_dir = (
            rag_settings.project_dir / ".rag_cache"
            if rag_settings is not None
            else None
        )
        resolved = (
            RoadTypeResolver(cache_dir=cache_dir).resolve(coordinates)
            if self.allow_network_road_lookup
            else RoadContext(notes=["Network road lookup disabled; using supplied context/data only."])
        )
        if rag_settings is not None:
            resolved = self._apply_reviewed_route_override(
                resolved, coordinates, rag_settings.project_dir
            )
        supplied = self.road_context_override or RoadContext()
        data_lanes = median_integer(frame.get("Expected Total Lanes", []))
        merged = resolved.model_copy(
            update={
                key: value
                for key, value in supplied.model_dump().items()
                if value not in (None, "", "unknown", [], {})
            }
        )
        if supplied.lanes_total is not None:
            merged = merged.model_copy(update={"lane_configuration_provisional": False})
            merged.notes.append(
                "The base road lane configuration was explicitly verified by the user/CLI; "
                "raw OSM lane counts remain recorded separately for traceability."
            )
        if supplied.carriageway == "unknown" and merged.carriageway == "unknown":
            inferred_carriageway = self._infer_carriageway_from_frame(frame, data_lanes)
            if inferred_carriageway != "unknown":
                merged = merged.model_copy(update={"carriageway": inferred_carriageway})
                merged.notes.append(
                    "Carriageway type was inferred from expected lanes and detected centre-line markings."
                )
        if supplied.lanes_total is None and data_lanes:
            if merged.carriageway in {"divided", "one_way"}:
                if merged.lanes_per_carriageway is None:
                    merged = merged.model_copy(
                        update={
                            "lanes_per_carriageway": data_lanes,
                            "lane_count_source": "input workbook (observed carriageway)",
                        }
                    )
                paired_lanes = merged.opposite_carriageway_lanes
                if (
                    merged.carriageway == "divided"
                    and merged.total_road_lanes is None
                    and merged.carriageway_count == 2
                    and merged.osm_way_lanes == data_lanes
                    and paired_lanes is not None
                    and paired_lanes >= data_lanes
                ):
                    merged = merged.model_copy(
                        update={
                            "lanes_total": data_lanes * 2,
                            "total_road_lanes": data_lanes * 2,
                            "lane_count_source": (
                                "provisional base through-lane count from workbook "
                                "and paired OSM carriageways"
                            ),
                            "lane_configuration_provisional": True,
                        }
                    )
                    merged.notes.append(
                        "The base through-lane total was inferred as twice the observed "
                        "carriageway lane count. A larger opposite OSM lanes=* value was "
                        "retained as a raw count because it may include an auxiliary/turn lane."
                    )
                else:
                    merged.notes.append(
                        "The workbook lane count was treated as lanes on the observed carriageway; "
                        "it was not doubled without a matched opposite carriageway."
                    )
            elif merged.total_road_lanes is None:
                merged = merged.model_copy(
                    update={
                        "lanes_total": data_lanes,
                        "total_road_lanes": data_lanes,
                        "carriageway_count": 1 if merged.carriageway == "undivided" else None,
                        "lane_count_source": "input workbook (road total)",
                        "lane_configuration_provisional": True,
                    }
                )
                merged.notes.append("Total lane count came from the input metrics file.")
        if (
            supplied.design_speed_kmph is None
            and merged.design_speed_kmph is None
            and merged.posted_speed_kmph is None
        ):
            detected_speed = self._infer_speed_from_sign_labels(frame)
            if detected_speed is not None:
                merged = merged.model_copy(
                    update={
                        "posted_speed_kmph": detected_speed,
                        "posted_speed_source": "numbered speed-limit sign detected in input workbook",
                    }
                )
                merged.notes.append(
                    "A numbered speed-limit sign in the input workbook supplied a posted-speed proxy; "
                    "it was not treated as verified design speed."
                )
        if not supplied.sign_class and "Traffic Sign Class" in frame.columns:
            sign_classes = {
                str(value).strip()
                for value in frame["Traffic Sign Class"].dropna().tolist()
                if str(value).strip() and str(value).strip().casefold() != "nan"
            }
            if len(sign_classes) == 1:
                merged = merged.model_copy(update={"sign_class": next(iter(sign_classes))})
                merged.notes.append("Traffic-sign class came from the input metrics file.")
            elif len(sign_classes) > 1:
                merged.notes.append(
                    "Multiple traffic-sign classes occur in the workbook; face-size rules require per-class review."
                )
            if supplied.sign_mounting == "unknown" and sign_classes and all(
                self._is_project_roadside_sign(label) for label in sign_classes
            ):
                merged = merged.model_copy(update={"sign_mounting": "shoulder"})
                merged.notes.append(
                    "Sign mounting was inferred as shoulder/roadside because this project detects only "
                    "roadside chevrons, warning signs, and roadside mandatory speed-limit signs."
                )
        return merged

    @staticmethod
    def _apply_reviewed_route_override(
        context: RoadContext,
        coordinates: list[tuple[float, float]],
        project_dir: Path,
    ) -> RoadContext:
        """Apply a narrow, traceable road-configuration override.

        Overrides never replace raw OSM fields. They only supply the reviewed
        base through-lane configuration used for standards applicability.
        """

        path = project_dir / "config" / "road_context_overrides.json"
        if not coordinates or not path.exists():
            return context
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            context.notes.append(
                f"Road-context override file could not be read: {path}."
            )
            return context

        latitudes = sorted(point[0] for point in coordinates)
        longitudes = sorted(point[1] for point in coordinates)
        centre = (
            latitudes[len(latitudes) // 2],
            longitudes[len(longitudes) // 2],
        )
        actual_ref = re.sub(r"\s+", "", context.highway_ref or "").casefold()
        for item in payload.get("overrides", []):
            if not isinstance(item, dict):
                continue
            expected_ref = re.sub(r"\s+", "", str(item.get("highway_ref", ""))).casefold()
            if not actual_ref or not expected_ref or actual_ref != expected_ref:
                continue
            bounds = item.get("bounds")
            if not isinstance(bounds, dict):
                continue
            try:
                inside = (
                    float(bounds["min_lat"]) <= centre[0] <= float(bounds["max_lat"])
                    and float(bounds["min_lon"]) <= centre[1] <= float(bounds["max_lon"])
                )
            except (KeyError, TypeError, ValueError):
                continue
            if not inside:
                continue
            review_status = str(item.get("review_status", "reviewer_verified")).casefold()
            source_label = (
                "reviewed route override"
                if review_status == "reviewer_verified"
                else "project route configuration"
            )
            update = {
                "carriageway": item.get("carriageway", context.carriageway),
                "lanes_per_carriageway": item.get("lanes_per_carriageway"),
                "carriageway_count": item.get("carriageway_count"),
                "total_road_lanes": item.get("total_road_lanes"),
                "lanes_total": item.get("total_road_lanes"),
                "lane_count_source": (
                    f"{source_label}: {item.get('id', 'unnamed')}"
                ),
                "lane_configuration_provisional": review_status != "reviewer_verified",
                "barrier_type": item.get("barrier_type", context.barrier_type),
            }
            notes = list(context.notes)
            notes.append(
                "A route-scoped lane/barrier configuration was applied from "
                f"{path.name} ({item.get('id', 'unnamed')}). Raw OSM lane counts were preserved."
            )
            source_note = str(item.get("source_note", "")).strip()
            if source_note:
                notes.append(source_note)
            return RoadContext.model_validate(
                {**context.model_dump(mode="python"), **update, "notes": notes}
            )
        return context

    @staticmethod
    def _infer_carriageway_from_frame(frame, expected_lanes: int | None) -> str:
        if not expected_lanes or expected_lanes < 2:
            return "unknown"
        centre_lines = median_integer(frame.get("Centre Lanes Detected", []))
        shoulder_lines = median_integer(frame.get("Shoulder Lanes Detected", []))
        if centre_lines is not None and centre_lines >= expected_lanes - 1:
            return "undivided"
        if centre_lines == 0 and shoulder_lines and shoulder_lines >= 2:
            return "divided"
        return "unknown"

    @staticmethod
    def _is_project_roadside_sign(label: str) -> bool:
        normalized = re.sub(r"[^a-z0-9]+", " ", label.casefold()).strip()
        return any(
            token in normalized
            for token in ("chevron", "warning sign", "mandatory sign", "speed limit")
        )

    @staticmethod
    def _infer_speed_from_sign_labels(frame) -> float | None:
        if "Traffic Sign Class" not in frame.columns:
            return None
        speeds: list[float] = []
        for value in frame["Traffic Sign Class"].dropna().tolist():
            match = re.search(
                r"(?:speed\s*limit|maximum\s*speed|mandatory\s*speed)\D{0,8}(\d{2,3})\b",
                str(value),
                re.IGNORECASE,
            )
            if match:
                speed = float(match.group(1))
                if 10 <= speed <= 180:
                    speeds.append(speed)
        if not speeds:
            return None
        return float(median_integer(speeds) or speeds[0])

    def _audit_rows(self, frame, threshold_set: ThresholdSet):
        import pandas as pd

        results: list[dict[str, object]] = []
        rules = threshold_set.results
        for index, row in frame.iterrows():
            coordinate = parse_coordinate_text(row.get("Raw OCR Text", ""))
            checks = {
                "Lane_Width": evaluate_measurements(
                    extract_numbers(row.get("Tracked Lane Widths (m)")), rules["min_lane_width"]
                ),
                "Radius_of_Curvature": evaluate_measurements(
                    extract_numbers(row.get("Radius of Curvature (m)")), rules["min_radius_curvature"]
                ),
                "Kerb_Height": evaluate_measurements(
                    extract_numbers(row.get("kerb_height")), rules["min_kerb_height"]
                ),
                "WBeam_Barrier_Height": evaluate_measurements(
                    extract_numbers(row.get("WBeam_crash barrier_height")),
                    rules["min_w_beam_barrier_height"],
                ),
                "Concrete_Barrier_Height": evaluate_measurements(
                    extract_numbers(row.get("Concrete_crash_barrier_height")),
                    rules["min_concrete_barrier_height"],
                ),
                "Traffic_Sign_Mounting_Height": evaluate_measurements(
                    extract_numbers(row.get("Height")), rules["min_sign_height"]
                ),
            }
            failures = [name for name, check in checks.items() if check.failed]
            reviews = [name for name, check in checks.items() if check.review_required]
            if failures:
                disposition = "NON_COMPLIANT"
            elif reviews:
                disposition = "REVIEW_REQUIRED"
            else:
                disposition = "COMPLIANT_ON_OBSERVED_PARAMETERS"
            result: dict[str, object] = {
                "Source_Row": index + 2,
                "Timestamp_s": row.get("Timestamp (sec)", index),
                "Latitude": coordinate[0] if coordinate else None,
                "Longitude": coordinate[1] if coordinate else None,
                "Road_Context": threshold_set.road_context.compact_description(),
                "Expected_Total_Lanes": row.get("Expected Total Lanes"),
                "Centre_Lanes_Detected": row.get("Centre Lanes Detected"),
                "Shoulder_Lanes_Detected": row.get("Shoulder Lanes Detected"),
                "Observed_Lane_Integrity": row.get("Lane Integrity Status"),
                "Observed_Lane_Widths_m": row.get("Tracked Lane Widths (m)"),
                "Observed_Radius_m": row.get("Radius of Curvature (m)"),
                "Observed_Kerb_Height_m": row.get("kerb_height"),
                "Observed_WBeam_Heights_m": row.get("WBeam_crash barrier_height"),
                "Observed_Concrete_Barrier_Heights_m": row.get(
                    "Concrete_crash_barrier_height"
                ),
                "Traffic_Sign_ID": row.get("ID"),
                "Traffic_Sign_Class": row.get("Traffic Sign Class"),
                "Observed_Sign_Mounting_Height_m": row.get("Height"),
                "Observed_Sign_Distance_m": row.get("Distance"),
                "Disposition": disposition,
                "Severity": self._heuristic_severity(len(failures), len(reviews)),
                "Recommendation": self._recommendation(
                    failures, reviews, checks, threshold_set
                ),
            }
            for name, check in checks.items():
                result[f"{name}_Status"] = check.status
                result[f"{name}_Evidence"] = check.detail
            results.append(result)
        return pd.DataFrame(results)

    @staticmethod
    def _recommendation(
        failures: list[str],
        reviews: list[str],
        checks: dict[str, CheckResult],
        threshold_set: ThresholdSet,
    ) -> str:
        if failures:
            actions = [
                RoadSafetyAuditPipeline._corrective_action(
                    name.replace("_", " "),
                    checks[name].detail,
                    CHECK_METRIC_KEYS.get(name),
                    threshold_set,
                )
                for name in failures
            ]
            return " ".join(actions)
        if reviews:
            labels = ", ".join(name.replace("_", " ") for name in reviews)
            return f"Do not issue a compliance conclusion for {labels}; complete the missing road context/source verification first."
        return "Observed parameters with applicable verified rules pass; retain the measurements and citations in the audit record."

    @staticmethod
    def _heuristic_severity(failure_count: int, review_count: int) -> str:
        """Map failed-category counts to presentation labels, not engineering risk."""

        if failure_count >= 3:
            return "HIGH SEVERITY (DANGEROUS)"
        if failure_count == 2:
            return "MEDIUM SEVERITY"
        if failure_count == 1:
            return "LOW SEVERITY"
        if review_count:
            return "REVIEW REQUIRED"
        return "SAFE"

    @staticmethod
    def _standards_rows(threshold_set: ThresholdSet) -> list[dict[str, object]]:
        standards_rows: list[dict[str, object]] = []
        for result in threshold_set.results.values():
            standards_rows.append(
                {
                    "Metric": result.metric_name,
                    "Key": result.metric_key,
                    "Status": result.status.value,
                    "Value_m": result.value_m,
                    "Second_Value_m": result.second_value_m,
                    "Comparator": result.comparator,
                    "Evidence_Quality_Score": result.evidence_quality_score,
                    "Calibrated_Correctness_Probability": (
                        result.calibrated_correctness_probability
                    ),
                    "Calibration_ID": result.calibration_id,
                    "Applicability_Basis": result.applicability_basis,
                    "Provisional": "YES" if result.provisional else "NO",
                    "Missing_Context": ", ".join(result.missing_context),
                    "Conditions": "; ".join(result.conditions),
                    "Standard": result.citation.standard_id if result.citation else None,
                    "Edition": result.citation.edition_year if result.citation else None,
                    "Source": result.citation.source if result.citation else None,
                    "Page": result.citation.page if result.citation else None,
                    "Section": result.citation.section if result.citation else None,
                    "Evidence_Quote": result.citation.quote if result.citation else None,
                    "Reason": result.reason,
                }
            )
        return standards_rows

    @staticmethod
    def _write_workbook(audited, threshold_set: ThresholdSet, output: Path) -> None:
        import pandas as pd
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        standards_rows = RoadSafetyAuditPipeline._standards_rows(threshold_set)
        metadata_rows = [
            ("Run ID", threshold_set.run_id),
            ("Generated UTC", datetime.now(timezone.utc).isoformat()),
            ("Collection", threshold_set.collection_name),
            ("Embedding model", threshold_set.embedding_model),
            ("Ollama model", threshold_set.llm_model),
            ("Road context", threshold_set.road_context.compact_description()),
            ("Road context JSON", json.dumps(threshold_set.road_context.model_dump(mode="json"))),
            ("Warnings", "; ".join(threshold_set.warnings)),
            (
                "Evidence-quality scoring",
                "Heuristic diagnostic only; it is not a calibrated probability and is not an audit-readiness gate.",
            ),
            (
                "Calibration status",
                "Not calibrated: no version-matched reviewed gold calibration artifact is loaded.",
            ),
            (
                "Severity model",
                "Count-based presentation heuristic only: 1/2/3+ failed categories map to low/medium/high; not an auditor-approved engineering risk model.",
            ),
        ]
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            audited.to_excel(writer, sheet_name="Audit Results", index=False)
            pd.DataFrame(standards_rows).to_excel(writer, sheet_name="Applicable Standards", index=False)
            pd.DataFrame(metadata_rows, columns=["Field", "Value"]).to_excel(
                writer, sheet_name="Run Metadata", index=False
            )
            for worksheet in writer.book.worksheets:
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.alignment = Alignment(wrap_text=True)
                for column_cells in worksheet.columns:
                    letter = get_column_letter(column_cells[0].column)
                    max_length = min(
                        60,
                        max(len(str(cell.value or "")) for cell in column_cells) + 2,
                    )
                    worksheet.column_dimensions[letter].width = max(10, max_length)
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _write_dashboard_html(
        audited, threshold_set: ThresholdSet, output: Path, source: Path
    ) -> None:
        """Write a portable, dependency-free audit dashboard with escaped data."""

        import pandas as pd

        counts = audited["Disposition"].value_counts().to_dict()
        total = len(audited)
        compliant = int(counts.get("COMPLIANT_ON_OBSERVED_PARAMETERS", 0))
        non_compliant = int(counts.get("NON_COMPLIANT", 0))
        review = int(counts.get("REVIEW_REQUIRED", 0))
        geotagged = int(audited["Latitude"].notna().sum())
        standards = pd.DataFrame(RoadSafetyAuditPipeline._standards_rows(threshold_set))

        status_columns = [column for column in audited.columns if column.endswith("_Status")]
        report_rows = audited.copy()
        report_rows["Checks_Requiring_Attention"] = report_rows.apply(
            lambda row: ", ".join(
                column.removesuffix("_Status").replace("_", " ")
                for column in status_columns
                if row.get(column) in {"FAIL", "REVIEW_REQUIRED"}
            ),
            axis=1,
        )
        report_rows["Location"] = report_rows.apply(
            lambda row: (
                f"{row['Latitude']:.6f}, {row['Longitude']:.6f}"
                if pd.notna(row.get("Latitude")) and pd.notna(row.get("Longitude"))
                else "Not available"
            ),
            axis=1,
        )
        display_columns = [
            "Source_Row",
            "Timestamp_s",
            "Location",
            "Disposition",
            "Severity",
            "Checks_Requiring_Attention",
            "Observed_Lane_Widths_m",
            "Observed_Radius_m",
            "Traffic_Sign_Class",
            "Recommendation",
        ]
        observations_table = report_rows[display_columns].to_html(
            index=False,
            border=0,
            classes="data-table",
            table_id="audit-table",
            escape=True,
            na_rep="",
        )
        standards_columns = [
            "Metric",
            "Status",
            "Value_m",
            "Comparator",
            "Evidence_Quality_Score",
            "Calibrated_Correctness_Probability",
            "Calibration_ID",
            "Applicability_Basis",
            "Provisional",
            "Standard",
            "Edition",
            "Page",
            "Section",
            "Evidence_Quote",
            "Reason",
        ]
        standards_table = standards[standards_columns].to_html(
            index=False,
            border=0,
            classes="data-table",
            table_id="standards-table",
            escape=True,
            na_rep="",
        )
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        context = escape(threshold_set.road_context.compact_description())
        warning_text = escape("; ".join(threshold_set.warnings) or "None")
        source_text = escape(str(source))
        run_id = escape(threshold_set.run_id)

        def proportion(value: int) -> float:
            return round(100 * value / max(total, 1), 1)

        html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Road Safety Audit Report</title>
<style>
:root {{--navy:#123047;--blue:#1f6f9f;--green:#18794e;--red:#b42318;--amber:#a15c00;--ink:#182230;--muted:#667085;--line:#d0d5dd;--paper:#fff;--bg:#f2f5f8;}}
* {{box-sizing:border-box}} body {{margin:0;background:var(--bg);color:var(--ink);font-family:Segoe UI,Arial,sans-serif;line-height:1.45}}
header {{background:linear-gradient(120deg,var(--navy),var(--blue));color:#fff;padding:34px max(24px,5vw)}}
header h1 {{margin:0 0 8px;font-size:clamp(26px,4vw,42px)}} header p {{margin:4px 0;opacity:.92}}
main {{max-width:1500px;margin:auto;padding:24px}} .panel {{background:var(--paper);border:1px solid var(--line);border-radius:12px;padding:20px;margin-bottom:20px;box-shadow:0 2px 8px #1018280d}}
.cards {{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:14px;margin-bottom:20px}}
.card {{background:#fff;border:1px solid var(--line);border-radius:12px;padding:16px}} .card strong {{display:block;font-size:30px}} .card span {{color:var(--muted)}}
.good strong {{color:var(--green)}} .bad strong {{color:var(--red)}} .review strong {{color:var(--amber)}}
.bar {{height:16px;border-radius:10px;overflow:hidden;background:#e4e7ec;display:flex;margin-top:12px}} .bar span {{height:100%}}
.bar .pass {{background:var(--green);width:{proportion(compliant)}%}} .bar .fail {{background:var(--red);width:{proportion(non_compliant)}%}} .bar .rev {{background:#f5a623;width:{proportion(review)}%}}
.toolbar {{display:flex;gap:12px;flex-wrap:wrap;margin:12px 0}} input {{min-width:280px;flex:1;padding:10px 12px;border:1px solid var(--line);border-radius:8px}}
button {{border:0;border-radius:8px;background:var(--blue);color:#fff;padding:10px 15px;cursor:pointer}}
.table-wrap {{overflow:auto;max-height:680px;border:1px solid var(--line);border-radius:9px}} .data-table {{border-collapse:collapse;width:100%;font-size:13px;background:#fff}}
.data-table th {{position:sticky;top:0;background:var(--navy);color:#fff;text-align:left;z-index:1}} .data-table th,.data-table td {{padding:9px 10px;border-bottom:1px solid #e4e7ec;vertical-align:top;min-width:95px}}
.data-table tbody tr:nth-child(even) {{background:#f8fafc}} .status-pass {{color:var(--green);font-weight:700}} .status-fail {{color:var(--red);font-weight:700}} .status-review {{color:var(--amber);font-weight:700}}
.meta {{display:grid;grid-template-columns:minmax(130px,220px) 1fr;gap:8px 18px}} .meta b {{color:var(--muted)}} h2 {{margin-top:0}} footer {{padding:24px;text-align:center;color:var(--muted)}}
@media print {{body {{background:#fff}} header {{background:#fff;color:#000;padding:12px}} main {{max-width:none;padding:0}} .panel,.card {{box-shadow:none;break-inside:avoid}} .toolbar {{display:none}} .table-wrap {{max-height:none;overflow:visible}} .data-table th {{position:static;background:#ddd;color:#000}}}}
</style>
</head>
<body>
<header><h1>Road Safety Audit Report</h1><p>Evidence-bound IRC threshold evaluation</p><p>Generated {generated}</p></header>
<main>
<section class="cards">
<div class="card"><strong>{total}</strong><span>Total observations</span></div>
<div class="card good"><strong>{compliant}</strong><span>Compliant on observed parameters</span></div>
<div class="card bad"><strong>{non_compliant}</strong><span>Non-compliant</span></div>
<div class="card review"><strong>{review}</strong><span>Review required</span></div>
<div class="card"><strong>{geotagged}</strong><span>Geotagged observations</span></div>
</section>
<section class="panel"><h2>Disposition overview</h2><div class="bar"><span class="pass"></span><span class="fail"></span><span class="rev"></span></div></section>
<section class="panel"><h2>Run information</h2><div class="meta"><b>Input file</b><span>{source_text}</span><b>Run ID</b><span>{run_id}</span><b>Road context</b><span>{context}</span><b>Embedding model</b><span>{escape(threshold_set.embedding_model)}</span><b>Ollama model</b><span>{escape(threshold_set.llm_model)}</span><b>Warnings</b><span>{warning_text}</span></div></section>
<section class="panel"><h2>Applicable standards and evidence</h2><div class="table-wrap">{standards_table}</div></section>
<section class="panel"><h2>Observation-level audit</h2><div class="toolbar"><input id="audit-search" placeholder="Filter rows, status, feature or recommendation"><button onclick="window.print()">Print / Save PDF</button></div><div class="table-wrap">{observations_table}</div></section>
</main>
<footer>Decision-support report. Confirm standards editions and findings with a qualified road-safety auditor.</footer>
<script>
const search=document.getElementById('audit-search'); const rows=[...document.querySelectorAll('#audit-table tbody tr')];
search.addEventListener('input',()=>{{const q=search.value.toLowerCase(); rows.forEach(r=>r.hidden=!r.innerText.toLowerCase().includes(q));}});
document.querySelectorAll('td').forEach(c=>{{const t=c.textContent.trim(); if(t==='PASS'||t==='COMPLIANT_ON_OBSERVED_PARAMETERS'||t==='found')c.classList.add('status-pass'); if(t==='FAIL'||t==='NON_COMPLIANT'||t==='invalid_evidence')c.classList.add('status-fail'); if(t==='REVIEW_REQUIRED'||t==='ambiguous'||t==='needs_context'||t==='not_found')c.classList.add('status-review');}});
</script>
</body></html>"""
        output.write_text(html, encoding="utf-8")

    @staticmethod
    def _format_measurement(
        value: float | None,
        rule: ThresholdResult,
        prefix: str = "",
    ) -> tuple[str, str]:
        if value is None:
            return "Absent", "ABSENT"
        rendered = f"{value:g}m"
        if not rule.audit_ready:
            label = rule.status.value.replace("_", " ").upper()
            if rule.value_m is not None and rule.comparator is not None:
                if rule.comparator == "range" and rule.second_value_m is not None:
                    low, high = sorted((rule.value_m, rule.second_value_m))
                    candidate = f"candidate {low:g}-{high:g}m"
                else:
                    candidate = f"candidate {rule.comparator} {rule.value_m:g}m"
                return (
                    f"{prefix}{rendered} ({candidate}; {label}) - REVIEW REQUIRED",
                    "REVIEW_REQUIRED",
                )
            return f"{prefix}{rendered} (threshold {label}) - REVIEW REQUIRED", "REVIEW_REQUIRED"
        check = evaluate_measurements([value], rule)
        assert rule.value_m is not None and rule.comparator is not None
        if rule.comparator == "range" and rule.second_value_m is not None:
            low, high = sorted((rule.value_m, rule.second_value_m))
            relation = f"between {low:g}m and {high:g}m"
        elif check.status == "PASS":
            relation = f"{rule.comparator} {rule.value_m:g}m"
        else:
            inverse = {">=": "<", ">": "<=", "<=": ">", "<": ">=", "=": "!="}
            relation = f"{inverse.get(rule.comparator, rule.comparator)} {rule.value_m:g}m"
        provisional = ""
        if rule.provisional:
            basis = rule.applicability_basis or "posted-speed proxy"
            provisional = f" (PROVISIONAL: {basis})"
        return (
            f"{prefix}{rendered} ({relation}) - {check.status}{provisional}",
            check.status,
        )

    @staticmethod
    def _format_measurements(
        values: list[float],
        rule: ThresholdResult,
        label: str,
    ) -> tuple[str, str]:
        """Evaluate and retain every measurement in a report cell.

        Overall precedence is deliberately deterministic: any failed value
        makes the feature fail; otherwise a non-audit-ready threshold makes
        every supplied value require review; only all passing values produce
        PASS. No value is silently discarded.
        """

        if not values:
            return "Absent", "ABSENT"
        multiple = len(values) > 1
        details: list[str] = []
        states: list[str] = []
        for position, value in enumerate(values, start=1):
            item_label = f"{label} {position}" if multiple else label
            detail, state = RoadSafetyAuditPipeline._format_measurement(
                value, rule, f"{item_label}: "
            )
            details.append(detail)
            states.append(state)
        if "FAIL" in states:
            overall = "FAIL"
        elif "REVIEW_REQUIRED" in states:
            overall = "REVIEW_REQUIRED"
        else:
            overall = "PASS"
        return " | ".join(details), overall

    @staticmethod
    def _lane_integrity(row, expected_lanes: int | None) -> str:
        if not expected_lanes:
            return "REVIEW REQUIRED: expected lane count unavailable"
        center_values = extract_numbers(row.get("Centre Lanes Detected"))
        shoulder_values = extract_numbers(row.get("Shoulder Lanes Detected"))
        if not center_values or not shoulder_values:
            return "REVIEW REQUIRED: detected lane-marking counts unavailable"
        detected_center = int(center_values[0])
        detected_shoulder = int(shoulder_values[0])
        missing_center = max(expected_lanes - 1 - detected_center, 0)
        missing_shoulder = max(2 - detected_shoulder, 0)
        issues: list[str] = []
        if missing_center:
            issues.append(f"{missing_center} centre line(s) missing")
        if missing_shoulder:
            issues.append(f"{missing_shoulder} shoulder line(s) missing")
        return "Intact" if not issues else " | ".join(issues)

    @staticmethod
    def _row_recommendation(
        failures: list[tuple[str, str, str | None]],
        reviews: list[tuple[str, str, str | None]],
        threshold_set: ThresholdSet,
    ) -> str:
        if not failures and not reviews:
            return (
                "Observed parameters with applicable verified rules pass. Retain the "
                "measurements and cited evidence in the audit record."
            )
        parts: list[str] = []
        if failures:
            parts.extend(
                RoadSafetyAuditPipeline._corrective_action(
                    name, detail, key, threshold_set
                )
                for name, detail, key in failures
            )
        if reviews:
            names = ", ".join(name for name, _detail, _key in reviews)
            parts.append(
                f"Do not issue a final compliance conclusion for {names} until the "
                "standard edition/applicability context is verified."
            )
        cited: list[str] = []
        for _name, _detail, key in [*failures, *reviews]:
            if not key or key not in threshold_set.results:
                continue
            citation = threshold_set.results[key].citation
            if citation and citation.label not in cited:
                cited.append(citation.label)
        if cited:
            parts.append("Evidence: " + "; ".join(cited) + ".")
        return " ".join(parts)

    @staticmethod
    def _corrective_action(
        name: str,
        detail: str,
        metric_key: str | None,
        threshold_set: ThresholdSet,
    ) -> str:
        """Create deterministic, measurement-specific corrective guidance."""

        if not metric_key or metric_key not in threshold_set.results:
            if "lane-marking" in name.casefold() or "line" in detail.casefold():
                return f"Restore the missing lane marking(s), then verify continuity: {detail}."
            return f"Rectify {name.casefold()} and remeasure: {detail}."

        rule = threshold_set.results[metric_key]
        if not rule.audit_ready or rule.value_m is None or rule.comparator is None:
            return f"Review {name.casefold()} before specifying corrective work: {detail}."

        measurements: list[tuple[str, float]] = []
        for segment in detail.split("|"):
            if "FAIL" not in segment.upper():
                continue
            match = re.match(
                r"\s*([^:|]+):\s*([-+]?\d+(?:\.\d+)?)m\b",
                segment,
                re.IGNORECASE,
            )
            if match:
                measurements.append((match.group(1).strip(), float(match.group(2))))

        low = rule.value_m
        high = rule.second_value_m

        def passes(value: float) -> bool:
            if rule.comparator == ">=":
                return value >= low
            if rule.comparator == ">":
                return value > low
            if rule.comparator == "<=":
                return value <= low
            if rule.comparator == "<":
                return value < low
            if rule.comparator == "range" and high is not None:
                lower, upper = sorted((low, high))
                return lower <= value <= upper
            return False

        if not measurements:
            observed_text = detail.split(";", 1)[0]
            measurements = [
                (name, value)
                for value in extract_numbers(observed_text)
                if not passes(value)
            ]
        if not measurements:
            return f"Rectify {name.casefold()} to the cited requirement and remeasure."

        def number(value: float) -> str:
            return f"{value:.3f}".rstrip("0").rstrip(".")

        actions: list[str] = []
        for label, value in measurements:
            if metric_key == "min_radius_curvature":
                shortfall = max(low - value, 0.0)
                actions.append(
                    f"Refer {label.casefold()} for geometric-design review: the measured "
                    f"radius is {number(value)} m, {number(shortfall)} m below the "
                    f"{number(low)} m requirement; consider curve redesign or an "
                    "engineer-approved speed-management treatment."
                )
                continue

            if rule.comparator == "range" and high is not None:
                lower, upper = sorted((low, high))
                if value < lower:
                    verb = (
                        "Raise"
                        if "height" in name.casefold() or "barrier" in name.casefold()
                        else "Increase"
                    )
                    delta = lower - value
                    actions.append(
                        f"{verb} {label} by at least {number(delta)} m, from "
                        f"{number(value)} m into the acceptable {number(lower)}-"
                        f"{number(upper)} m range, then remeasure."
                    )
                elif value > upper:
                    verb = (
                        "Lower"
                        if "height" in name.casefold() or "barrier" in name.casefold()
                        else "Reduce"
                    )
                    delta = value - upper
                    actions.append(
                        f"{verb} {label} by at least {number(delta)} m, from "
                        f"{number(value)} m into the acceptable {number(lower)}-"
                        f"{number(upper)} m range, then remeasure."
                    )
                continue

            if rule.comparator in {">=", ">"}:
                delta = max(low - value, 0.0)
                if metric_key == "min_lane_width":
                    verb, subject = "Widen", label
                elif metric_key == "min_sign_height":
                    verb = "Raise"
                    subject = "the traffic sign" if label == "Sign height" else label
                else:
                    verb, subject = "Increase", label
                target = "at least" if rule.comparator == ">=" else "above"
                actions.append(
                    f"{verb} {subject} by at least {number(delta)} m, from "
                    f"{number(value)} m to {target} {number(low)} m, then remeasure."
                )
                continue

            if rule.comparator in {"<=", "<"}:
                delta = max(value - low, 0.0)
                target = "at most" if rule.comparator == "<=" else "below"
                actions.append(
                    f"Reduce {label} by at least {number(delta)} m, from "
                    f"{number(value)} m to {target} {number(low)} m, then remeasure."
                )

        return " ".join(actions) or f"Rectify {name.casefold()} and remeasure."

    def _audit_rows(self, frame, threshold_set: ThresholdSet):
        """Create the requested 16-column DL/CV-to-audit report layout."""

        import pandas as pd

        rules = threshold_set.results
        results: list[dict[str, object]] = []
        road_type = threshold_set.road_context.road_class or "UNVERIFIED ROAD CLASS"
        context_lanes = (
            threshold_set.road_context.lanes_per_carriageway
            if threshold_set.road_context.carriageway in {"divided", "one_way"}
            else threshold_set.road_context.total_road_lanes
        )
        for index, row in frame.iterrows():
            coordinate = parse_coordinate_text(row.get("Raw OCR Text", ""))
            expected_values = extract_numbers(row.get("Expected Total Lanes"))
            # The CV row observes one carriageway. Prefer its own expected-lane
            # field, then fall back to the context count for that carriageway.
            expected_lanes = int(expected_values[0]) if expected_values else context_lanes
            lane_integrity = self._lane_integrity(row, expected_lanes)
            failures: list[tuple[str, str, str | None]] = []
            reviews: list[tuple[str, str, str | None]] = []

            radius_values = extract_numbers(row.get("Radius of Curvature (m)"))
            radius_status, radius_state = self._format_measurements(
                radius_values,
                rules["min_radius_curvature"],
                "Radius",
            )

            lane_values = extract_numbers(row.get("Tracked Lane Widths (m)"))
            if not lane_values:
                lane_status = "Absent"
                lane_overall = "Absent"
            elif not rules["min_lane_width"].audit_ready:
                lane_status = " | ".join(
                    f"Lane {position}: {value:g}m (REVIEW REQUIRED)"
                    for position, value in enumerate(lane_values, start=1)
                )
                lane_overall = "REVIEW_REQUIRED"
            else:
                lane_parts: list[str] = []
                lane_states: list[str] = []
                for position, value in enumerate(lane_values, start=1):
                    checked = evaluate_measurements([value], rules["min_lane_width"])
                    lane_parts.append(f"Lane {position}: {value:g}m ({checked.status})")
                    lane_states.append(checked.status)
                lane_status = " | ".join(lane_parts)
                lane_overall = "PASS" if all(state == "PASS" for state in lane_states) else "FAIL"

            kerb_values = extract_numbers(row.get("kerb_height"))
            kerb_status, kerb_state = self._format_measurements(
                kerb_values,
                rules["min_kerb_height"],
                "Kerb",
            )
            w_beam_values = extract_numbers(row.get("WBeam_crash barrier_height"))
            concrete_values = extract_numbers(row.get("Concrete_crash_barrier_height"))
            generic_barrier_values = extract_numbers(row.get("Crash barrier_height"))
            if generic_barrier_values and not w_beam_values and not concrete_values:
                if threshold_set.road_context.barrier_type == "w_beam":
                    w_beam_values = generic_barrier_values
                elif threshold_set.road_context.barrier_type == "concrete":
                    concrete_values = generic_barrier_values
            barrier_parts: list[str] = []
            barrier_states: list[str] = []
            if w_beam_values:
                detail, state = self._format_measurements(
                    w_beam_values,
                    rules["min_w_beam_barrier_height"],
                    "W-Beam",
                )
                barrier_parts.append(detail)
                barrier_states.append(state)
            if concrete_values:
                detail, state = self._format_measurements(
                    concrete_values,
                    rules["min_concrete_barrier_height"],
                    "Concrete",
                )
                barrier_parts.append(detail)
                barrier_states.append(state)
            generic_untyped = bool(
                generic_barrier_values and not w_beam_values and not concrete_values
            )
            if generic_untyped:
                rendered = ", ".join(f"{value:g}m" for value in generic_barrier_values)
                barrier_status = (
                    f"Barrier: {rendered} (type is unavailable; cannot select W-beam or "
                    "concrete rule) - REVIEW REQUIRED"
                )
                barrier_states.append("REVIEW_REQUIRED")
            else:
                barrier_status = " | ".join(barrier_parts) if barrier_parts else "Absent"

            sign_values = extract_numbers(row.get("Height"))
            sign_status, sign_state = self._format_measurements(
                sign_values,
                rules["min_sign_height"],
                "Sign height",
            )

            checks = [
                ("Radius of curvature", radius_status, radius_state, "min_radius_curvature"),
                ("Lane width", lane_status, lane_overall, "min_lane_width"),
                ("Kerb height", kerb_status, kerb_state, "min_kerb_height"),
                (
                    "Crash barrier height",
                    barrier_status,
                    "FAIL" if "FAIL" in barrier_states else (
                        "REVIEW_REQUIRED" if "REVIEW_REQUIRED" in barrier_states else "PASS"
                    ) if barrier_states else "ABSENT",
                    "min_w_beam_barrier_height" if w_beam_values else "min_concrete_barrier_height",
                ),
                ("Traffic-sign mounting height", sign_status, sign_state, "min_sign_height"),
            ]
            if lane_integrity not in {"Intact"}:
                state = "REVIEW_REQUIRED" if lane_integrity.startswith("REVIEW REQUIRED") else "FAIL"
                checks.append(("Lane-marking integrity", lane_integrity, state, None))
            for name, detail, state, key in checks:
                if state == "FAIL":
                    failures.append((name, detail, key))
                elif state == "REVIEW_REQUIRED":
                    reviews.append((name, detail, key))

            severity = self._heuristic_severity(len(failures), len(reviews))

            results.append(
                {
                    "Timestamp_s": row.get("Timestamp (sec)", index),
                    "Latitude": coordinate[0] if coordinate else None,
                    "Longitude": coordinate[1] if coordinate else None,
                    "Detected_Road_Type": road_type,
                    "Expected_Lanes": expected_lanes,
                    "Lane_Integrity": lane_integrity,
                    "Kerb_Available": "Yes" if kerb_values else "Absent",
                    "Crash_Barrier_Available": (
                        "Yes"
                        if w_beam_values or concrete_values or generic_barrier_values
                        else "Absent"
                    ),
                    "Radius_of_Curvature_Status": radius_status,
                    "Lane_Width_Status": lane_status,
                    "Overall_Lane_Width_Status": lane_overall,
                    "Kerb_Height_Status": kerb_status,
                    "Crash_Barrier_Status": barrier_status,
                    "Traffic_Sign_Height_Status": sign_status,
                    "Overall_Severity": severity,
                    "Professional_Suggestions": self._row_recommendation(
                        failures, reviews, threshold_set
                    ),
                }
            )
        return pd.DataFrame(results)

    @staticmethod
    def _write_html(
        audited, threshold_set: ThresholdSet, output: Path, source: Path
    ) -> None:
        """Write an interactive Leaflet map with disclosed heuristic severity."""


        map_rows = audited.dropna(subset=["Latitude", "Longitude"]).copy()
        records = json.loads(map_rows.to_json(orient="records"))
        marker_json = json.dumps(records, ensure_ascii=False).replace("</", "<\\/")
        if records:
            start_lat = float(records[0]["Latitude"])
            start_lon = float(records[0]["Longitude"])
        else:
            start_lat, start_lon = 20.5937, 78.9629
        # Legend totals describe only geotagged observations visible on the map.
        severity_counts = map_rows["Overall_Severity"].value_counts().to_dict()
        standards_items: list[str] = []
        for result in threshold_set.results.values():
            if result.value_m is not None and result.comparator is not None:
                if result.comparator == "range" and result.second_value_m is not None:
                    low, high = sorted((result.value_m, result.second_value_m))
                    value = f"{low:g}-{high:g} m"
                else:
                    value = f"{result.comparator} {result.value_m:g} m"
                if not result.audit_ready:
                    value = f"Candidate {value} - {result.status.value.replace('_', ' ').upper()}"
            else:
                value = result.status.value.replace("_", " ").upper()
            citation = escape(result.citation.label) if result.citation else escape(result.reason)
            basis = (
                f"<br><small><b>Basis:</b> {escape(result.applicability_basis)}"
                f"{' — PROVISIONAL' if result.provisional else ''}</small>"
                if result.applicability_basis
                else ""
            )
            quality = (
                f"<br><small><b>Evidence quality score:</b> "
                f"{result.evidence_quality_score:.2f} "
                "(heuristic diagnostic; not a probability)</small>"
            )
            if result.calibrated_correctness_probability is None:
                calibration = (
                    "<br><small><b>Calibrated correctness probability:</b> "
                    "not available — reviewed gold calibration is not loaded</small>"
                )
            else:
                calibration = (
                    f"<br><small><b>Calibrated correctness probability:</b> "
                    f"{result.calibrated_correctness_probability:.1%} "
                    f"({escape(result.calibration_id or 'unversioned calibration')})</small>"
                )
            standards_items.append(
                f"<li><b>{escape(result.metric_name)}:</b> {escape(value)}"
                f"<br><small>{citation}</small>{basis}{quality}{calibration}</li>"
            )
        standards_html = "".join(standards_items)
        context = escape(threshold_set.road_context.compact_description())
        source_name = escape(source.name)
        warning_text = escape("; ".join(threshold_set.warnings) or "None")
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        html = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Road Safety Audit Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" crossorigin="">
<style>
html,body,#map{{height:100%;margin:0}} body{{font-family:Arial,sans-serif}} #map{{z-index:1}}
.box{{position:fixed;z-index:9999;background:#fff;border:2px solid #667085;border-radius:8px;box-shadow:2px 2px 9px #0004;padding:10px;font-size:13px}}
.legend{{top:20px;right:20px;width:150px;line-height:1.7}} .summary{{left:20px;top:20px;width:260px}}
.standards{{left:20px;bottom:20px;width:min(430px,calc(100vw - 40px));max-height:34vh;overflow:auto}}
.standards summary{{cursor:pointer;font-weight:bold}} .standards ul{{padding-left:20px}} .standards li{{margin:8px 0}}
.dot{{font-size:20px;vertical-align:-2px}} .popup{{width:310px;line-height:1.45}} .popup hr{{margin:8px 0;border:0;border-top:1px solid #ddd}}
@media(max-width:700px){{.summary{{display:none}}.legend{{top:10px;right:10px}}.standards{{left:10px;bottom:10px;width:calc(100vw - 20px);max-height:26vh}}}}
</style></head><body>
<div id="map"></div>
<div class="box summary"><b>Evidence-bound Road Safety Audit</b><br>Input: {source_name}<br>Generated: {generated}<br>Observations: {len(audited)}<br>Geotagged: {len(records)}<br><small>{context}</small><br><small><b>Audit mode:</b> {warning_text}</small><br><small><b>Network notice:</b> this map loads Leaflet and OpenStreetMap tiles from the internet. Opening it may disclose the viewed map area to those providers.</small></div>
<div class="box legend"><b>Heuristic Severity Legend</b><br><span class="dot" style="color:green">●</span> SAFE ({severity_counts.get('SAFE', 0)})<br><span class="dot" style="color:blue">●</span> LOW SEVERITY ({severity_counts.get('LOW SEVERITY', 0)})<br><span class="dot" style="color:orange">●</span> MEDIUM SEVERITY ({severity_counts.get('MEDIUM SEVERITY', 0)})<br><span class="dot" style="color:red">●</span> HIGH SEVERITY / DANGEROUS ({severity_counts.get('HIGH SEVERITY (DANGEROUS)', 0)})<br><span class="dot" style="color:purple">●</span> REVIEW REQUIRED ({severity_counts.get('REVIEW REQUIRED', 0)})<br><small>Count-based presentation only (1/2/3+ failed categories); not an auditor-approved engineering risk model.</small></div>
<details class="box standards"><summary>Applicable standards and evidence</summary><ul>{standards_html}</ul><small>Decision support only; confirm editions and applicability with a qualified road-safety auditor.</small></details>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" crossorigin=""></script>
<script>
const observations={marker_json};
const map=L.map('map').setView([{start_lat},{start_lon}],16);
L.tileLayer('https://tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',{{maxZoom:19,attribution:'&copy; OpenStreetMap contributors'}}).addTo(map);
const colorFor=s=>s==='SAFE'?'green':s==='LOW SEVERITY'?'blue':s==='MEDIUM SEVERITY'?'orange':s.startsWith('HIGH')?'red':'purple';
const e=v=>String(v??'').replace(/[&<>\"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}}[c]));
const bounds=[];
for(const row of observations){{
 const color=colorFor(row.Overall_Severity); bounds.push([row.Latitude,row.Longitude]);
 const sign=row.Traffic_Sign_Height_Status!=='Absent'?`<b>Traffic Sign:</b> ${{e(row.Traffic_Sign_Height_Status)}}<br>`:'';
 const popup=`<div class="popup"><b>Severity:</b> <span style="color:${{color}};font-weight:bold">${{e(row.Overall_Severity)}}</span><br><b>Timestamp:</b> ${{e(row.Timestamp_s)}} s<br><b>Coordinates:</b> ${{Number(row.Latitude).toFixed(6)}}, ${{Number(row.Longitude).toFixed(6)}}<br><b>Road Type:</b> ${{e(row.Detected_Road_Type)}}<br><b>Lane Integrity:</b> ${{e(row.Lane_Integrity)}}<br><b>Radius Check:</b> ${{e(row.Radius_of_Curvature_Status)}}<br><b>Lanes Check:</b> ${{e(row.Lane_Width_Status)}}<br><b>Kerb Check:</b> ${{e(row.Kerb_Height_Status)}}<br><b>Barrier Check:</b> ${{e(row.Crash_Barrier_Status)}}<br>${{sign}}<hr><b>Recommendation:</b> ${{e(row.Professional_Suggestions)}}</div>`;
 L.circleMarker([row.Latitude,row.Longitude],{{radius:7,color,fill:true,fillColor:color,fillOpacity:.8,weight:3}}).addTo(map).bindPopup(popup,{{maxWidth:360}});
}}
if(bounds.length>1)map.fitBounds(bounds,{{padding:[30,30],maxZoom:17}});
</script></body></html>"""
        output.write_text(html, encoding="utf-8")
